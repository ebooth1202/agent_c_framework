"""
Workbook Manager for Excel Tools.

Handles workbook lifecycle operations: creation, loading, saving, and metadata management.
"""

import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from ..models import WorkbookMetadata, SheetInfo, LoadResult, SaveResult, OperationResult


class WorkbookManager:
    """
    Manages Excel workbook lifecycle operations.
    
    Handles creation, loading, saving, and metadata tracking of Excel workbooks
    with proper separation from the tool interface.
    """
    
    def __init__(self):
        """Initialize the workbook manager."""
        self.current_workbook: Optional[Workbook] = None
        self.current_workbook_path: Optional[str] = None
        self.current_workbook_metadata: Optional[WorkbookMetadata] = None
    
    def _generate_operation_id(self) -> str:
        """Generate a unique operation ID for tracking."""
        return f"excel_op_{int(time.time() * 1000)}_{uuid.uuid4().hex[:8]}"
    
    def _gather_sheet_info(self, workbook: Workbook, read_only: bool = False) -> tuple[list[SheetInfo], int]:
        """
        Gather information about all sheets in a workbook.
        
        Args:
            workbook: The workbook to analyze
            read_only: Whether the workbook was opened read-only
            
        Returns:
            Tuple of (sheet_info_list, total_rows)
        """
        sheets_info = []
        total_rows = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row if not read_only else 0
            max_col = sheet.max_column if not read_only else 0
            is_active = sheet == workbook.active
            
            sheet_info = SheetInfo(
                name=sheet_name,
                max_row=max_row,
                max_column=max_col,
                is_active=is_active
            )
            
            sheets_info.append(sheet_info)
            total_rows += max_row
        
        return sheets_info, total_rows
    
    def create_workbook(self) -> OperationResult:
        """
        Create a new Excel workbook in memory.
        
        Returns:
            OperationResult with success status and workbook information
        """
        try:
            operation_id = self._generate_operation_id()
            
            # Create new workbook
            self.current_workbook = Workbook()
            self.current_workbook_path = None
            
            # Create metadata
            self.current_workbook_metadata = WorkbookMetadata(
                operation_id=operation_id,
                created_at=datetime.now(),
                sheets=[{'name': 'Sheet', 'max_row': 0, 'max_column': 0}],
                total_rows=0
            )
            
            return OperationResult(
                success=True,
                message='New Excel workbook created in memory',
                operation_id=operation_id,
                data={
                    'sheets': self.current_workbook.sheetnames,
                    'default_sheet': self.current_workbook.active.title
                }
            )
            
        except Exception as e:
            return OperationResult(
                success=False,
                message='Failed to create workbook',
                error=str(e)
            )
    
    def load_workbook(self, file_path: str, read_only: bool = False) -> LoadResult:
        """
        Load an Excel workbook from file system.
        
        Args:
            file_path: Absolute path to the Excel file
            read_only: Whether to open in read-only mode
            
        Returns:
            LoadResult with workbook information or error
        """
        try:
            operation_id = self._generate_operation_id()
            
            # Check if file exists
            if not Path(file_path).exists():
                return LoadResult(
                    success=False,
                    file_path=file_path,
                    sheets=[],
                    total_rows=0,
                    read_only=read_only,
                    error=f'File not found: {file_path}'
                )
            
            # Load workbook
            if read_only:
                self.current_workbook = load_workbook(file_path, read_only=True, data_only=True)
            else:
                self.current_workbook = load_workbook(file_path)
            
            self.current_workbook_path = file_path
            
            # Gather workbook information
            sheets_info, total_rows = self._gather_sheet_info(self.current_workbook, read_only)
            
            # Create metadata
            self.current_workbook_metadata = WorkbookMetadata(
                operation_id=operation_id,
                loaded_at=datetime.now(),
                file_path=file_path,
                read_only=read_only,
                sheets=[sheet.to_dict() for sheet in sheets_info],
                total_rows=total_rows
            )
            
            return LoadResult(
                success=True,
                file_path=file_path,
                sheets=sheets_info,
                total_rows=total_rows,
                read_only=read_only,
                operation_id=operation_id
            )
            
        except Exception as e:
            return LoadResult(
                success=False,
                file_path=file_path,
                sheets=[],
                total_rows=0,
                read_only=read_only,
                error=f'Error loading workbook: {str(e)}'
            )
    
    def save_workbook(self, file_path: str) -> tuple[SaveResult, Optional[bytes]]:
        """
        Save the current workbook to bytes and return save information.
        
        Args:
            file_path: Path where the workbook will be saved
            
        Returns:
            Tuple of (SaveResult, workbook_bytes)
        """
        try:
            if not self.current_workbook:
                return SaveResult(
                    success=False,
                    file_path=file_path,
                    file_size_bytes=0,
                    sheets_count=0,
                    error='No workbook is currently loaded'
                ), None
            
            operation_id = self._generate_operation_id()
            
            # Save workbook to bytes
            workbook_bytes = save_virtual_workbook(self.current_workbook)
            
            # Update metadata
            self.current_workbook_path = file_path
            if self.current_workbook_metadata:
                self.current_workbook_metadata.saved_at = datetime.now()
                self.current_workbook_metadata.file_path = file_path
            
            return SaveResult(
                success=True,
                file_path=file_path,
                file_size_bytes=len(workbook_bytes),
                sheets_count=len(self.current_workbook.sheetnames),
                operation_id=operation_id
            ), workbook_bytes
            
        except Exception as e:
            return SaveResult(
                success=False,
                file_path=file_path,
                file_size_bytes=0,
                sheets_count=0,
                error=f'Error saving workbook: {str(e)}'
            ), None
    
    def get_workbook_info(self) -> OperationResult:
        """
        Get information about the current workbook.
        
        Returns:
            OperationResult with workbook information
        """
        if not self.current_workbook:
            return OperationResult(
                success=False,
                message='No workbook is currently loaded',
                error='No workbook available'
            )
        
        sheets_info, _ = self._gather_sheet_info(self.current_workbook)
        
        return OperationResult(
            success=True,
            message='Workbook information retrieved',
            data={
                'sheets': [sheet.to_dict() for sheet in sheets_info],
                'total_sheets': len(sheets_info),
                'active_sheet': self.current_workbook.active.title,
                'metadata': self.current_workbook_metadata.to_dict() if self.current_workbook_metadata else {}
            }
        )
    
    def create_sheet(self, sheet_name: str, index: Optional[int] = None) -> OperationResult:
        """
        Create a new sheet in the current workbook.
        
        Args:
            sheet_name: Name for the new sheet
            index: Position where to insert the sheet (0-based)
            
        Returns:
            OperationResult with creation status
        """
        try:
            if not self.current_workbook:
                return OperationResult(
                    success=False,
                    message='No workbook is currently loaded',
                    error='No workbook available'
                )
            
            # Check if sheet already exists
            if sheet_name in self.current_workbook.sheetnames:
                return OperationResult(
                    success=False,
                    message=f'Sheet "{sheet_name}" already exists',
                    error=f'Sheet name conflict: {sheet_name}'
                )
            
            # Create the sheet
            if index is not None:
                new_sheet = self.current_workbook.create_sheet(sheet_name, index)
            else:
                new_sheet = self.current_workbook.create_sheet(sheet_name)
            
            return OperationResult(
                success=True,
                message=f'Sheet "{sheet_name}" created successfully',
                data={
                    'sheet_name': sheet_name,
                    'sheet_index': self.current_workbook.sheetnames.index(sheet_name),
                    'total_sheets': len(self.current_workbook.sheetnames)
                }
            )
            
        except Exception as e:
            return OperationResult(
                success=False,
                message=f'Failed to create sheet "{sheet_name}"',
                error=str(e)
            )
    
    def has_workbook(self) -> bool:
        """Check if a workbook is currently loaded."""
        return self.current_workbook is not None
    
    def get_workbook(self) -> Optional[Workbook]:
        """Get the current workbook instance."""
        return self.current_workbook
    
    def get_workbook_path(self) -> Optional[str]:
        """Get the current workbook file path."""
        return self.current_workbook_path
    
    def get_workbook_metadata(self) -> Optional[WorkbookMetadata]:
        """Get the current workbook metadata."""
        return self.current_workbook_metadata