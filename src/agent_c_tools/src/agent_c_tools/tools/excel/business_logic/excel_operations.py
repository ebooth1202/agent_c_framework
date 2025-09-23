"""
Excel Operations Manager for Excel Tools.

Handles core Excel read and write operations using openpyxl.
"""

import asyncio
import threading
import time
import uuid
from typing import Any, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from ..models import WriteResult, ReadResult, OperationResult


class ExcelOperations:
    """
    Handles core Excel read and write operations.
    
    Provides methods for reading from and writing to Excel worksheets
    with proper error handling and progress reporting.
    """
    
    DEFAULT_CHUNK_SIZE = 10000
    
    def __init__(self):
        """Initialize the Excel operations manager."""
        pass
    
    def _generate_operation_id(self) -> str:
        """Generate a unique operation ID for tracking."""
        return f"excel_op_{int(time.time() * 1000)}_{uuid.uuid4().hex[:8]}"
    
    async def get_next_available_row(self, workbook: Workbook, sheet_name: str) -> int:
        """
        Get the next available row for writing in a sheet.
        
        Args:
            workbook: The workbook to check
            sheet_name: Name of the sheet
            
        Returns:
            Next available row number (1-based)
        """
        if sheet_name not in workbook.sheetnames:
            return 1
            
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        
        # Check if the last row actually has data
        if max_row == 1:
            # Check if row 1 has data
            has_data = any(sheet.cell(1, col).value is not None for col in range(1, sheet.max_column + 1))
            return 2 if has_data else 1
        
        return max_row + 1
    
    async def append_records(
        self,
        workbook: Workbook,
        records: List[List[str]],
        sheet_name: str = 'Sheet',
        headers: Optional[List[str]] = None,
        agent_id: Optional[str] = None,
        chunk_size: int = DEFAULT_CHUNK_SIZE,
        client_wants_cancel: Optional[threading.Event] = None,
        progress_callback: Optional[callable] = None
    ) -> WriteResult:
        """
        Append records to the end of a sheet.
        
        Args:
            workbook: The workbook to write to
            records: List of records to append
            sheet_name: Name of the sheet to append to
            headers: Optional column headers to add if sheet is empty
            agent_id: Unique identifier for the agent performing the operation
            chunk_size: Number of records to process in each batch
            client_wants_cancel: Event to check for cancellation
            progress_callback: Optional callback for progress updates
            
        Returns:
            WriteResult with operation details
        """
        try:
            if not records:
                return WriteResult(
                    success=False,
                    records_written=0,
                    start_row=0,
                    end_row=0,
                    sheet_name=sheet_name,
                    error='No records provided to append'
                )
            
            operation_id = self._generate_operation_id()
            
            if not agent_id:
                agent_id = f'agent_{uuid.uuid4().hex[:8]}'
            
            # Ensure sheet exists
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            
            sheet = workbook[sheet_name]
            
            # Add headers if sheet is empty and headers provided
            start_row = await self.get_next_available_row(workbook, sheet_name)
            headers_added = False
            
            if start_row == 1 and headers:
                for col_idx, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col_idx, value=header)
                start_row = 2
                headers_added = True
            
            # Process records in chunks for memory efficiency
            total_written = 0
            current_row = start_row
            
            for chunk_start in range(0, len(records), chunk_size):
                if client_wants_cancel and client_wants_cancel.is_set():
                    break
                
                chunk_end = min(chunk_start + chunk_size, len(records))
                chunk_records = records[chunk_start:chunk_end]
                
                # Write chunk to sheet
                for record in chunk_records:
                    for col_idx, value in enumerate(record, 1):
                        sheet.cell(row=current_row, column=col_idx, value=value)
                    current_row += 1
                    total_written += 1
                
                # Update progress for large operations
                if progress_callback and len(records) > chunk_size and chunk_end < len(records):
                    progress = (chunk_end / len(records)) * 100
                    await progress_callback(f'Appending records: {progress:.1f}% complete ({chunk_end}/{len(records)})')
            
            end_row = current_row - 1
            
            return WriteResult(
                success=True,
                records_written=total_written,
                start_row=start_row,
                end_row=end_row,
                sheet_name=sheet_name,
                operation_id=operation_id,
                agent_id=agent_id,
                headers_added=headers_added
            )
            
        except Exception as e:
            return WriteResult(
                success=False,
                records_written=0,
                start_row=0,
                end_row=0,
                sheet_name=sheet_name,
                agent_id=agent_id,
                error=f'Error appending records: {str(e)}'
            )
    
    async def write_to_reserved_rows(
        self,
        workbook: Workbook,
        records: List[List[str]],
        sheet_name: str,
        start_row: int,
        max_rows: int
    ) -> WriteResult:
        """
        Write data to a specific row range (for reserved rows).
        
        Args:
            workbook: The workbook to write to
            records: List of records to write
            sheet_name: Name of the sheet
            start_row: Starting row number (1-based)
            max_rows: Maximum number of rows available
            
        Returns:
            WriteResult with operation details
        """
        try:
            if len(records) > max_rows:
                return WriteResult(
                    success=False,
                    records_written=0,
                    start_row=start_row,
                    end_row=start_row,
                    sheet_name=sheet_name,
                    error=f'Too many records ({len(records)}) for available space ({max_rows} rows)'
                )
            
            # Get sheet
            sheet = workbook[sheet_name]
            
            # Write records to specific rows
            current_row = start_row
            for record in records:
                for col_idx, value in enumerate(record, 1):
                    sheet.cell(row=current_row, column=col_idx, value=value)
                current_row += 1
            
            end_row = start_row + len(records) - 1
            
            return WriteResult(
                success=True,
                records_written=len(records),
                start_row=start_row,
                end_row=end_row,
                sheet_name=sheet_name
            )
            
        except Exception as e:
            return WriteResult(
                success=False,
                records_written=0,
                start_row=start_row,
                end_row=start_row,
                sheet_name=sheet_name,
                error=f'Error writing to reserved rows: {str(e)}'
            )
    
    def read_sheet_data(
        self,
        workbook: Workbook,
        sheet_name: str = 'Sheet',
        start_row: int = 1,
        end_row: Optional[int] = None,
        start_column: str = 'A',
        end_column: Optional[str] = None,
        include_headers: bool = False,
        max_rows: int = 10000
    ) -> ReadResult:
        """
        Read data from a sheet with optional range specification.
        
        Args:
            workbook: The workbook to read from
            sheet_name: Name of the sheet to read from
            start_row: Starting row (1-based)
            end_row: Ending row (1-based), None for last row with data
            start_column: Starting column (e.g., "A")
            end_column: Ending column (e.g., "Z"), None for last column with data
            include_headers: Whether to treat first row as headers
            max_rows: Maximum number of rows to read
            
        Returns:
            ReadResult with sheet data
        """
        try:
            # Check if sheet exists
            if sheet_name not in workbook.sheetnames:
                return ReadResult(
                    success=False,
                    error=f'Sheet "{sheet_name}" does not exist'
                )
            
            sheet = workbook[sheet_name]
            
            # Determine actual range
            if not end_row:
                end_row = min(sheet.max_row, start_row + max_rows - 1)
            
            if not end_column:
                end_column = get_column_letter(sheet.max_column)
            
            # Convert column letters to indices
            start_col_idx = column_index_from_string(start_column)
            end_col_idx = column_index_from_string(end_column)
            
            # Read data
            data = []
            headers = []
            
            for row_idx, row in enumerate(sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col_idx,
                max_col=end_col_idx,
                values_only=True
            )):
                if row_idx == 0 and include_headers:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                else:
                    data.append([str(cell) if cell is not None else '' for cell in row])
            
            return ReadResult(
                success=True,
                data=data,
                headers=headers if include_headers else None,
                rows_read=len(data),
                columns_read=end_col_idx - start_col_idx + 1,
                range_info=f'{start_column}{start_row}:{end_column}{end_row}'
            )
            
        except Exception as e:
            return ReadResult(
                success=False,
                error=f'Error reading sheet data: {str(e)}'
            )
    
    def get_next_available_row_info(
        self, 
        workbook: Workbook, 
        sheet_name: str,
        reserved_rows: int = 0
    ) -> OperationResult:
        """
        Get information about the next available row in a sheet.
        
        Args:
            workbook: The workbook to check
            sheet_name: Name of the sheet
            reserved_rows: Number of rows currently reserved
            
        Returns:
            OperationResult with row information
        """
        try:
            # Check if sheet exists
            if sheet_name not in workbook.sheetnames:
                return OperationResult(
                    success=True,
                    message=f'Sheet "{sheet_name}" does not exist',
                    data={
                        'sheet_name': sheet_name,
                        'next_available_row': 1,
                        'reserved_rows': reserved_rows,
                        'next_safe_row': 1 + reserved_rows,
                        'sheet_exists': False
                    }
                )
            
            sheet = workbook[sheet_name]
            next_row = sheet.max_row + 1 if sheet.max_row > 0 else 1
            
            return OperationResult(
                success=True,
                message='Next available row determined',
                data={
                    'sheet_name': sheet_name,
                    'next_available_row': next_row,
                    'reserved_rows': reserved_rows,
                    'next_safe_row': next_row + reserved_rows,
                    'sheet_exists': True
                }
            )
            
        except Exception as e:
            return OperationResult(
                success=False,
                message='Error getting next available row',
                error=str(e)
            )