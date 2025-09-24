"""
Component tests for the Excel business logic functionality.
These tests call the business logic classes directly to validate their functionality.
"""
import pytest
import json
import sys
import os
from pathlib import Path
import tempfile

# Add parent directories to import business logic
current_dir = Path(__file__).parent
tool_dir = current_dir.parent
src_dir = tool_dir.parent.parent.parent  # Get to src level
sys.path.insert(0, str(src_dir))

from agent_c_tools.tools.excel.business_logic.workbook_manager import WorkbookManager
from agent_c_tools.tools.excel.business_logic.concurrency_manager import ConcurrencyManager
from agent_c_tools.tools.excel.business_logic.excel_operations import ExcelOperations
from agent_c_tools.tools.excel.models import (
    WorkbookMetadata, SheetInfo, LoadResult, SaveResult, OperationResult,
    WriteResult, ReadResult, ReservationInfo, ReservationStatus
)


class TestWorkbookManagerComponent:
    """Component tests for WorkbookManager that test real functionality."""
    
    def setup_method(self):
        """Set up test fixtures before each test method."""
        self.workbook_manager = WorkbookManager()
    
    @pytest.mark.component
    def test_workbook_manager_initialization(self):
        """Test WorkbookManager initializes correctly."""
        assert hasattr(self.workbook_manager, 'current_workbook')
        assert hasattr(self.workbook_manager, 'current_workbook_metadata')
        assert self.workbook_manager.current_workbook is None
        assert self.workbook_manager.current_workbook_metadata is None
    
    @pytest.mark.component
    def test_create_workbook_success(self):
        """Test successful workbook creation."""
        result = self.workbook_manager.create_workbook()
        
        # Validate it's a successful operation
        assert isinstance(result, OperationResult)
        assert result.success
        assert "created in memory" in result.message.lower()
        
        # Validate workbook was created
        assert self.workbook_manager.has_workbook()
        assert self.workbook_manager.current_workbook is not None
        
        # Validate metadata was initialized
        assert self.workbook_manager.current_workbook_metadata is not None
        assert isinstance(self.workbook_manager.current_workbook_metadata, WorkbookMetadata)
    
    @pytest.mark.component
    def test_get_workbook_info_with_workbook(self):
        """Test getting workbook info when workbook exists."""
        # Create a workbook first
        self.workbook_manager.create_workbook()
        
        result = self.workbook_manager.get_workbook_info()
        
        # Validate successful result
        assert isinstance(result, OperationResult)
        assert result.success
        assert result.data is not None
        assert "sheets" in result.data
        assert "total_sheets" in result.data
        
        # Validate sheets data
        sheets = result.data["sheets"]
        assert isinstance(sheets, list)
        assert len(sheets) > 0  # Should have at least one default sheet
    
    @pytest.mark.component
    def test_get_workbook_info_without_workbook(self):
        """Test getting workbook info when no workbook exists."""
        result = self.workbook_manager.get_workbook_info()
        
        # Should return error
        assert isinstance(result, OperationResult)
        assert not result.success
        assert "no workbook" in result.error.lower()
    
    @pytest.mark.component
    def test_create_sheet_success(self):
        """Test successful sheet creation."""
        # Create workbook first
        self.workbook_manager.create_workbook()
        
        # Create a new sheet
        sheet_name = "TestSheet"
        result = self.workbook_manager.create_sheet(sheet_name)
        
        # Validate successful creation
        assert isinstance(result, OperationResult)
        assert result.success
        assert sheet_name in result.message
        
        # Validate sheet exists in workbook
        workbook = self.workbook_manager.get_workbook()
        assert sheet_name in workbook.sheetnames
    
    @pytest.mark.component
    def test_create_duplicate_sheet_error(self):
        """Test error handling when creating duplicate sheet."""
        # Create workbook first
        self.workbook_manager.create_workbook()
        
        # Create a sheet
        sheet_name = "TestSheet"
        self.workbook_manager.create_sheet(sheet_name)
        
        # Try to create same sheet again
        result = self.workbook_manager.create_sheet(sheet_name)
        
        # Should return error
        assert isinstance(result, OperationResult)
        assert not result.success
        assert "name conflict" in result.error.lower()


class TestConcurrencyManagerComponent:
    """Component tests for ConcurrencyManager that test real functionality."""
    
    def setup_method(self):
        """Set up test fixtures before each test method."""
        self.concurrency_manager = ConcurrencyManager()
    
    @pytest.mark.component
    def test_concurrency_manager_initialization(self):
        """Test ConcurrencyManager initializes correctly."""
        assert hasattr(self.concurrency_manager, 'sheet_row_counts')
        assert hasattr(self.concurrency_manager, 'row_reservations')
        assert hasattr(self.concurrency_manager, '_write_lock')
        
        assert isinstance(self.concurrency_manager.sheet_row_counts, dict)
        assert isinstance(self.concurrency_manager.row_reservations, dict)
        assert len(self.concurrency_manager.sheet_row_counts) == 0
        assert len(self.concurrency_manager.row_reservations) == 0
    
    @pytest.mark.component
    @pytest.mark.asyncio
    async def test_reserve_rows_success(self):
        """Test successful row reservation."""
        result = await self.concurrency_manager.reserve_rows(
            row_count=10,
            sheet_name="TestSheet", 
            current_max_row=5,
            agent_id="test_agent"
        )
        
        # Validate successful reservation
        assert isinstance(result, OperationResult)
        assert result.success
        assert result.data is not None
        assert "reservation_id" in result.data
        assert "start_row" in result.data
        assert "end_row" in result.data
        assert "agent_id" in result.data
        
        # Validate reservation details
        assert result.data["agent_id"] == "test_agent"
        assert result.data["end_row"] - result.data["start_row"] + 1 == 10
    
    @pytest.mark.component
    @pytest.mark.asyncio
    async def test_multiple_reservations_no_overlap(self):
        """Test that multiple reservations don't overlap."""
        # First reservation
        result1 = await self.concurrency_manager.reserve_rows(
            row_count=5,
            sheet_name="TestSheet",
            current_max_row=5,
            agent_id="agent1"
        )
        
        # Second reservation
        result2 = await self.concurrency_manager.reserve_rows(
            row_count=3,
            sheet_name="TestSheet", 
            current_max_row=5,  # Same starting point
            agent_id="agent2"
        )
        
        # Both should succeed
        assert result1.success
        assert result2.success
        
        # Ranges should not overlap
        start1, end1 = result1.data["start_row"], result1.data["end_row"]
        start2, end2 = result2.data["start_row"], result2.data["end_row"]
        
        # Second reservation should start after first ends
        assert start2 > end1
        
        # Should have different reservation IDs
        assert result1.data["reservation_id"] != result2.data["reservation_id"]
    
    @pytest.mark.component 
    def test_get_operation_status(self):
        """Test getting operation status."""
        status = self.concurrency_manager.get_operation_status()
        
        # Should return a dictionary with status information
        assert isinstance(status, dict)
        assert "active_reservations" in status
        assert "row_tracking" in status
        assert isinstance(status["active_reservations"], dict)
        assert isinstance(status["row_tracking"], dict)


class TestExcelOperationsComponent:
    """Component tests for ExcelOperations that test real functionality."""
    
    def setup_method(self):
        """Set up test fixtures before each test method."""
        self.excel_operations = ExcelOperations()
        # Create a mock workbook for testing
        from openpyxl import Workbook
        self.test_workbook = Workbook()
        self.test_sheet = self.test_workbook.active
        self.test_sheet.title = "TestSheet"
    
    @pytest.mark.component
    def test_excel_operations_initialization(self):
        """Test ExcelOperations initializes correctly."""
        assert hasattr(self.excel_operations, 'logger')
        # Add other initialization checks as needed
    
    @pytest.mark.component
    def test_get_next_available_row_info_empty_sheet(self):
        """Test getting next available row info for empty sheet."""
        result = self.excel_operations.get_next_available_row_info(
            workbook=self.test_workbook,
            sheet_name="TestSheet", 
            reserved_rows=0
        )
        
        # Should return valid result
        assert isinstance(result, OperationResult)
        assert result.success
        assert result.data is not None
        assert "next_available_row" in result.data
        assert "sheet_name" in result.data
        
        # Empty sheet should start at row 1
        assert result.data["next_available_row"] >= 1
        assert result.data["sheet_name"] == "TestSheet"
    
    @pytest.mark.component
    def test_get_next_available_row_info_with_reserved_rows(self):
        """Test getting next available row info with reserved rows."""
        result = self.excel_operations.get_next_available_row_info(
            workbook=self.test_workbook,
            sheet_name="TestSheet",
            reserved_rows=5
        )
        
        # Should account for reserved rows
        assert result.success
        assert result.data["next_available_row"] >= 1  # Should account for reserved rows in next_safe_row
    
    @pytest.mark.component
    def test_read_sheet_data_empty_sheet(self):
        """Test reading data from empty sheet."""
        result = self.excel_operations.read_sheet_data(
            workbook=self.test_workbook,
            sheet_name="TestSheet",
            start_row=1,
            max_rows=100
        )
        
        # Should return valid but empty result
        assert isinstance(result, ReadResult)
        assert result.success
        assert isinstance(result.data, list)
        assert result.rows_read >= 0  # Could be 0 for empty sheet
    
    @pytest.mark.component
    def test_read_sheet_data_nonexistent_sheet(self):
        """Test reading data from non-existent sheet."""
        result = self.excel_operations.read_sheet_data(
            workbook=self.test_workbook,
            sheet_name="NonExistentSheet",
            start_row=1,
            max_rows=100
        )
        
        # Should return error
        assert isinstance(result, ReadResult)
        assert not result.success
        assert result.error is not None
        assert "not found" in result.error.lower() or "does not exist" in result.error.lower()


class TestModelsComponent:
    """Component tests for Excel models and data structures."""
    
    @pytest.mark.component
    def test_operation_result_to_dict(self):
        """Test OperationResult serialization."""
        result = OperationResult(
            success=True,
            message="Test successful",
            operation_id="test_123",
            data={"key": "value"}
        )
        
        result_dict = result.to_dict()
        
        assert isinstance(result_dict, dict)
        assert result_dict["success"] == True
        assert result_dict["message"] == "Test successful"
        assert result_dict["operation_id"] == "test_123"
        assert "key" in result_dict  # data should be merged into result
        assert result_dict["key"] == "value"
    
    @pytest.mark.component
    def test_write_result_to_dict(self):
        """Test WriteResult serialization."""
        result = WriteResult(
            success=True,
            records_written=100,
            start_row=1,
            end_row=100,
            sheet_name="TestSheet",
            operation_id="write_123"
        )
        
        result_dict = result.to_dict()
        
        assert result_dict["success"] == True
        assert result_dict["records_written"] == 100
        assert result_dict["start_row"] == 1
        assert result_dict["end_row"] == 100
        assert result_dict["sheet_name"] == "TestSheet"
        assert result_dict["operation_id"] == "write_123"
    
    @pytest.mark.component
    def test_read_result_to_dict(self):
        """Test ReadResult serialization."""
        test_data = [["Name", "Age"], ["John", "30"], ["Jane", "25"]]
        test_headers = ["Name", "Age"]
        
        result = ReadResult(
            success=True,
            data=test_data,
            headers=test_headers,
            rows_read=2
        )
        
        result_dict = result.to_dict()
        
        assert result_dict["success"] == True
        assert result_dict["rows_read"] == 2
        assert result_dict["data"] == test_data
        assert result_dict["headers"] == test_headers


if __name__ == "__main__":
    pytest.main([__file__, "-v"])