"""
Excel Tools for Agent C Framework

Provides comprehensive Excel file manipulation capabilities with support for:
- High-volume data writing (thousands to millions of records)
- Multi-agent concurrent operations with conflict resolution
- Memory-efficient processing for large datasets
- Bulk operations and batch processing
- Integration with other data processing toolsets
"""

import io
import json
import time
import uuid
import asyncio
import threading
from typing import Any, Dict, List, Optional, Union, Tuple
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.writer.excel import save_virtual_workbook

from agent_c.toolsets import Toolset, json_schema
from agent_c_tools.helpers.validate_kwargs import validate_required_fields
from agent_c_tools.helpers.path_helper import create_unc_path, ensure_file_extension, os_file_system_path
from agent_c_tools.helpers.media_file_html_helper import get_file_html
from agent_c_tools.tools.workspace.tool import WorkspaceTools


class ExcelTools(Toolset):
    """
    Gives your agent comprehensive Excel file manipulation capabilities designed for high-volume operations
    and multi-agent workflows. Your agent can create, read, write, and modify Excel files with support for
    thousands to millions of records, concurrent writing by multiple agents, and memory-efficient processing
    of large datasets from any data source.
    """
    
    DEFAULT_EXCEL_FOLDER = 'excel_data'
    MAX_EXCEL_TOKEN_SIZE = 35000
    MAX_ROWS_PER_SHEET = 1000000  # Stay under Excel's 1,048,576 limit
    DEFAULT_CHUNK_SIZE = 10000    # Process data in chunks for memory efficiency
    
    def __init__(self, **kwargs):
        """Initialize the ExcelTools."""
        super().__init__(**kwargs, name='excel')
        
        # Dependencies
        self.workspace_tool: Optional[WorkspaceTools] = None
        
        # Current workbook state
        self.current_workbook: Optional[Workbook] = None
        self.current_workbook_path: Optional[str] = None
        self.current_workbook_metadata: Dict[str, Any] = {}
        
        # Row tracking for concurrent operations
        self.row_reservations: Dict[str, Dict[str, Any]] = {}
        self.sheet_row_counts: Dict[str, int] = {}
        
        # Operation tracking
        self.active_operations: Dict[str, Dict[str, Any]] = {}
        
        # Concurrency control
        self._write_lock = asyncio.Lock()
        self._reservation_lock = asyncio.Lock()
        
    async def post_init(self):
        """Initialize dependencies after tool chest is available."""
        self.workspace_tool = self.tool_chest.available_tools.get('WorkspaceTools')
        if not self.workspace_tool:
            self.logger.error("WorkspaceTools dependency not available")

    def _generate_operation_id(self) -> str:
        """Generate a unique operation ID for tracking."""
        return f"excel_op_{int(time.time() * 1000)}_{uuid.uuid4().hex[:8]}"

    def _is_response_too_big(self, response: str) -> bool:
        """Check if response exceeds token limit."""
        if not hasattr(self.tool_chest, 'agent'):
            return len(response) > self.MAX_EXCEL_TOKEN_SIZE * 4  # Rough estimate
        
        token_count = self.tool_chest.agent.count_tokens(response)
        return token_count > self.MAX_EXCEL_TOKEN_SIZE

    async def _get_next_available_row(self, sheet_name: str) -> int:
        """Get the next available row for writing in a sheet."""
        if not self.current_workbook:
            return 1
            
        if sheet_name not in self.current_workbook.sheetnames:
            return 1
            
        sheet = self.current_workbook[sheet_name]
        # Find the last row with data
        max_row = sheet.max_row
        
        # Check if the last row actually has data
        if max_row == 1:
            # Check if row 1 has data
            has_data = any(sheet.cell(1, col).value is not None for col in range(1, sheet.max_column + 1))
            return 2 if has_data else 1
        
        return max_row + 1

    async def _update_sheet_row_count(self, sheet_name: str, new_count: int):
        """Update the tracked row count for a sheet."""
        self.sheet_row_counts[sheet_name] = max(
            self.sheet_row_counts.get(sheet_name, 0), 
            new_count
        )

    @json_schema(
        description='Create a new Excel workbook in memory',
        params={}
    )
    async def create_workbook(self, **kwargs) -> str:
        """
        Create a new Excel workbook in memory.
        
        Returns:
            str: JSON response with success status and workbook information
        """
        try:
            operation_id = self._generate_operation_id()
            self.logger.info(f"Creating new workbook (operation: {operation_id})")
            
            # Create new workbook
            self.current_workbook = Workbook()
            self.current_workbook_path = None
            self.current_workbook_metadata = {
                'created_at': datetime.now().isoformat(),
                'operation_id': operation_id,
                'sheets': ['Sheet'],  # Default sheet
                'total_rows': 0
            }
            
            # Reset tracking
            self.row_reservations.clear()
            self.sheet_row_counts.clear()
            
            return json.dumps({
                'success': True,
                'message': 'New Excel workbook created in memory',
                'operation_id': operation_id,
                'sheets': self.current_workbook.sheetnames,
                'default_sheet': self.current_workbook.active.title
            })
            
        except Exception as e:
            error_msg = f"Error creating workbook: {str(e)}"
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })

    @json_schema(
        description='Load an Excel workbook from workspace using UNC path',
        params={
            'path': {
                'type': 'string',
                'description': 'UNC-style path to the Excel file (e.g., //workspace/path/file.xlsx)',
                'required': True
            },
            'read_only': {
                'type': 'boolean', 
                'description': 'Open workbook in read-only mode for better performance with large files',
                'required': False,
                'default': False
            }
        }
    )
    async def load_workbook(self, **kwargs) -> str:
        """
        Load an Excel workbook from workspace.
        
        Args:
            **kwargs: Contains path and optional read_only flag
            
        Returns:
            str: JSON response with workbook information or error
        """
        try:
            # Validate required fields
            success, message = validate_required_fields(kwargs, ['path'])
            if not success:
                return json.dumps({'success': False, 'error': message})
            
            unc_path = kwargs.get('path')
            read_only = kwargs.get('read_only', False)
            operation_id = self._generate_operation_id()
            
            self.logger.info(f"Loading workbook from {unc_path} (operation: {operation_id})")
            
            # Validate workspace path
            error, workspace, relative_path = self.workspace_tool.validate_and_get_workspace_path(unc_path)
            if error:
                return json.dumps({'success': False, 'error': error})
            
            # Get OS path for openpyxl
            os_path = os_file_system_path(self.workspace_tool, unc_path)
            if not os_path:
                return json.dumps({'success': False, 'error': 'Could not resolve file path'})
            
            # Check if file exists
            if not Path(os_path).exists():
                return json.dumps({'success': False, 'error': f'File not found: {unc_path}'})
            
            # Load workbook
            if read_only:
                self.current_workbook = load_workbook(os_path, read_only=True, data_only=True)
            else:
                self.current_workbook = load_workbook(os_path)
            
            self.current_workbook_path = unc_path
            
            # Gather workbook metadata
            sheets_info = []
            total_rows = 0
            
            for sheet_name in self.current_workbook.sheetnames:
                sheet = self.current_workbook[sheet_name]
                max_row = sheet.max_row if not read_only else 0
                max_col = sheet.max_column if not read_only else 0
                
                sheets_info.append({
                    'name': sheet_name,
                    'max_row': max_row,
                    'max_column': max_col
                })
                
                total_rows += max_row
                self.sheet_row_counts[sheet_name] = max_row
            
            self.current_workbook_metadata = {
                'loaded_at': datetime.now().isoformat(),
                'operation_id': operation_id,
                'file_path': unc_path,
                'read_only': read_only,
                'sheets': sheets_info,
                'total_rows': total_rows
            }
            
            # Reset reservations for new workbook
            self.row_reservations.clear()
            
            return json.dumps({
                'success': True,
                'message': f'Excel workbook loaded successfully',
                'operation_id': operation_id,
                'file_path': unc_path,
                'sheets': sheets_info,
                'total_rows': total_rows,
                'read_only': read_only
            })
            
        except Exception as e:
            error_msg = f"Error loading workbook: {str(e)}"
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })

    @json_schema(
        description='Save the current workbook to workspace using UNC path',
        params={
            'path': {
                'type': 'string',
                'description': 'UNC-style path where to save the Excel file (e.g., //workspace/path/file.xlsx)',
                'required': True
            },
            'create_backup': {
                'type': 'boolean',
                'description': 'Create a backup of existing file before overwriting',
                'required': False,
                'default': False
            }
        }
    )
    async def save_workbook(self, **kwargs) -> str:
        """
        Save the current workbook to workspace.
        
        Args:
            **kwargs: Contains path and optional backup flag
            
        Returns:
            str: JSON response with save status
        """
        try:
            if not self.current_workbook:
                return json.dumps({
                    'success': False, 
                    'error': 'No workbook is currently loaded. Create or load a workbook first.'
                })
            
            # Validate required fields
            success, message = validate_required_fields(kwargs, ['path'])
            if not success:
                return json.dumps({'success': False, 'error': message})
            
            unc_path = kwargs.get('path')
            create_backup = kwargs.get('create_backup', False)
            operation_id = self._generate_operation_id()
            tool_context = kwargs.get('tool_context', {})
            
            self.logger.info(f"Saving workbook to {unc_path} (operation: {operation_id})")
            
            # Ensure .xlsx extension
            unc_path = ensure_file_extension(unc_path, 'xlsx')
            
            # Create backup if requested
            if create_backup:
                error, workspace, relative_path = self.workspace_tool.validate_and_get_workspace_path(unc_path)
                if not error:
                    os_path = os_file_system_path(self.workspace_tool, unc_path)
                    if os_path and Path(os_path).exists():
                        backup_path = unc_path.replace('.xlsx', f'_backup_{int(time.time())}.xlsx')
                        await self.workspace_tool.cp(src_path=unc_path, dest_path=backup_path)
                        self.logger.info(f"Backup created: {backup_path}")
            
            # Save workbook to bytes
            workbook_bytes = save_virtual_workbook(self.current_workbook)
            
            # Write to workspace
            result = await self.workspace_tool.internal_write_bytes(
                path=unc_path,
                data=workbook_bytes,
                mode='write'
            )
            
            # Check for write errors
            result_data = json.loads(result)
            if 'error' in result_data:
                return json.dumps({
                    'success': False,
                    'error': f"Failed to save workbook: {result_data['error']}"
                })
            
            # Update metadata
            self.current_workbook_path = unc_path
            self.current_workbook_metadata['saved_at'] = datetime.now().isoformat()
            self.current_workbook_metadata['file_path'] = unc_path
            
            # Get OS path for media event
            os_path = os_file_system_path(self.workspace_tool, unc_path)
            
            # Raise media event
            if os_path:
                await self._raise_render_media(
                    sent_by_class=self.__class__.__name__,
                    sent_by_function='save_workbook',
                    content_type="text/html",
                    content=get_file_html(
                        os_path=os_path,
                        unc_path=unc_path,
                        additional_html="Excel workbook saved successfully"
                    ),
                    tool_context=tool_context
                )
            
            return json.dumps({
                'success': True,
                'message': 'Excel workbook saved successfully',
                'operation_id': operation_id,
                'file_path': unc_path,
                'file_size_bytes': len(workbook_bytes),
                'sheets': len(self.current_workbook.sheetnames),
                'backup_created': create_backup
            })
            
        except Exception as e:
            error_msg = f"Error saving workbook: {str(e)}"
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })

    @json_schema(
        description='List all sheets in the current workbook with basic information',
        params={}
    )
    async def list_sheets(self, **kwargs) -> str:
        """
        List all sheets in the current workbook.
        
        Returns:
            str: JSON response with sheet information
        """
        try:
            if not self.current_workbook:
                return json.dumps({
                    'success': False,
                    'error': 'No workbook is currently loaded. Create or load a workbook first.'
                })
            
            sheets_info = []
            for sheet_name in self.current_workbook.sheetnames:
                sheet = self.current_workbook[sheet_name]
                
                # Get basic sheet info
                sheet_info = {
                    'name': sheet_name,
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'is_active': sheet == self.current_workbook.active
                }
                
                # Add tracked row count if available
                if sheet_name in self.sheet_row_counts:
                    sheet_info['tracked_row_count'] = self.sheet_row_counts[sheet_name]
                
                sheets_info.append(sheet_info)
            
            return json.dumps({
                'success': True,
                'sheets': sheets_info,
                'total_sheets': len(sheets_info),
                'active_sheet': self.current_workbook.active.title
            })
            
        except Exception as e:
            error_msg = f"Error listing sheets: {str(e)}"
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })

    @json_schema(
        description='Create a new sheet in the current workbook',
        params={
            'sheet_name': {
                'type': 'string',
                'description': 'Name for the new sheet',
                'required': True
            },
            'index': {
                'type': 'integer',
                'description': 'Position where to insert the sheet (0-based). If not provided, adds at the end.',
                'required': False
            }
        }
    )
    async def create_sheet(self, **kwargs) -> str:
        """
        Create a new sheet in the current workbook.
        
        Args:
            **kwargs: Contains sheet_name and optional index
            
        Returns:
            str: JSON response with creation status
        """
        try:
            if not self.current_workbook:
                return json.dumps({
                    'success': False,
                    'error': 'No workbook is currently loaded. Create or load a workbook first.'
                })
            
            # Validate required fields
            success, message = validate_required_fields(kwargs, ['sheet_name'])
            if not success:
                return json.dumps({'success': False, 'error': message})
            
            sheet_name = kwargs.get('sheet_name')
            index = kwargs.get('index')
            
            # Check if sheet already exists
            if sheet_name in self.current_workbook.sheetnames:
                return json.dumps({
                    'success': False,
                    'error': f'Sheet "{sheet_name}" already exists'
                })
            
            # Create the sheet
            if index is not None:
                new_sheet = self.current_workbook.create_sheet(sheet_name, index)
            else:
                new_sheet = self.current_workbook.create_sheet(sheet_name)
            
            # Initialize row tracking
            self.sheet_row_counts[sheet_name] = 0
            
            self.logger.info(f"Created sheet: {sheet_name}")
            
            return json.dumps({
                'success': True,
                'message': f'Sheet "{sheet_name}" created successfully',
                'sheet_name': sheet_name,
                'sheet_index': self.current_workbook.sheetnames.index(sheet_name),
                'total_sheets': len(self.current_workbook.sheetnames)
            })
            
        except Exception as e:
            error_msg = f"Error creating sheet: {str(e)}"
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })


    @json_schema(
        description='Reserve a range of rows for concurrent writing by multiple agents',
        params={
            'row_count': {
                'type': 'integer',
                'description': 'Number of rows to reserve',
                'required': True
            },
            'sheet_name': {
                'type': 'string',
                'description': 'Name of the sheet to reserve rows in',
                'required': False,
                'default': 'Sheet'
            },
            'agent_id': {
                'type': 'string',
                'description': 'Unique identifier for the agent requesting the reservation',
                'required': False
            }
        }
    )
    async def reserve_rows(self, **kwargs) -> str:
        """
        Reserve a range of rows for exclusive writing by an agent.
        This prevents conflicts when multiple agents write simultaneously.
        
        Args:
            **kwargs: Contains row_count, sheet_name, and optional agent_id
            
        Returns:
            str: JSON response with reservation details
        """
        try:
            if not self.current_workbook:
                return json.dumps({
                    'success': False,
                    'error': 'No workbook is currently loaded. Create or load a workbook first.'
                })
            
            # Validate required fields
            success, message = validate_required_fields(kwargs, ['row_count'])
            if not success:
                return json.dumps({'success': False, 'error': message})
            
            row_count = kwargs.get('row_count')
            sheet_name = kwargs.get('sheet_name', 'Sheet')
            agent_id = kwargs.get('agent_id', f'agent_{uuid.uuid4().hex[:8]}')
            
            if row_count <= 0:
                return json.dumps({
                    'success': False,
                    'error': 'row_count must be greater than 0'
                })
            
            async with self._reservation_lock:
                # Ensure sheet exists
                if sheet_name not in self.current_workbook.sheetnames:
                    self.current_workbook.create_sheet(sheet_name)
                    self.sheet_row_counts[sheet_name] = 0
                
                # Get next available row
                next_row = await self._get_next_available_row(sheet_name)
                
                # Account for any existing reservations
                reserved_rows = sum(
                    res['row_count'] for res in self.row_reservations.values()
                    if res['sheet_name'] == sheet_name
                )
                
                start_row = max(next_row, self.sheet_row_counts.get(sheet_name, 0) + 1) + reserved_rows
                end_row = start_row + row_count - 1
                
                # Create reservation
                reservation_id = f'res_{int(time.time() * 1000)}_{uuid.uuid4().hex[:8]}'
                
                self.row_reservations[reservation_id] = {
                    'agent_id': agent_id,
                    'sheet_name': sheet_name,
                    'start_row': start_row,
                    'end_row': end_row,
                    'row_count': row_count,
                    'reserved_at': datetime.now().isoformat(),
                    'status': 'active'
                }
                
                self.logger.info(f'Reserved rows {start_row}-{end_row} in sheet "{sheet_name}" for agent {agent_id}')
                
                return json.dumps({
                    'success': True,
                    'message': f'Reserved {row_count} rows in sheet "{sheet_name}"',
                    'reservation_id': reservation_id,
                    'agent_id': agent_id,
                    'sheet_name': sheet_name,
                    'start_row': start_row,
                    'end_row': end_row,
                    'row_count': row_count
                })
                
        except Exception as e:
            error_msg = f'Error reserving rows: {str(e)}'
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })

    @json_schema(
        description='Append records to the end of a sheet - safe for concurrent operations',
        params={
            'records': {
                'type': 'array',
                'description': 'Array of records to append. Each record should be an array of values.',
                'required': True,
                'items': {
                    'type': 'array',
                    'items': {'type': 'string'}
                }
            },
            'sheet_name': {
                'type': 'string',
                'description': 'Name of the sheet to append to',
                'required': False,
                'default': 'Sheet'
            },
            'headers': {
                'type': 'array',
                'description': 'Column headers to add if sheet is empty',
                'required': False,
                'items': {'type': 'string'}
            },
            'agent_id': {
                'type': 'string',
                'description': 'Unique identifier for the agent performing the operation',
                'required': False
            },
            'chunk_size': {
                'type': 'integer',
                'description': 'Number of records to process in each batch for memory efficiency',
                'required': False,
                'default': 10000
            }
        }
    )
    async def append_records(self, **kwargs) -> str:
        """
        Append records to the end of a sheet. This operation is safe for concurrent use
        as each agent appends to the current end of the sheet.
        
        Args:
            **kwargs: Contains records, sheet_name, headers, agent_id, chunk_size
            
        Returns:
            str: JSON response with append results
        """
        try:
            if not self.current_workbook:
                return json.dumps({
                    'success': False,
                    'error': 'No workbook is currently loaded. Create or load a workbook first.'
                })
            
            # Validate required fields
            success, message = validate_required_fields(kwargs, ['records'])
            if not success:
                return json.dumps({'success': False, 'error': message})
            
            records = kwargs.get('records')
            sheet_name = kwargs.get('sheet_name', 'Sheet')
            headers = kwargs.get('headers')
            agent_id = kwargs.get('agent_id', f'agent_{uuid.uuid4().hex[:8]}')
            chunk_size = kwargs.get('chunk_size', self.DEFAULT_CHUNK_SIZE)
            tool_context = kwargs.get('tool_context', {})
            client_wants_cancel: threading.Event = tool_context.get('client_wants_cancel', threading.Event())
            
            if not records:
                return json.dumps({
                    'success': False,
                    'error': 'No records provided to append'
                })
            
            operation_id = self._generate_operation_id()
            self.logger.info(f'Starting append operation {operation_id} for agent {agent_id}: {len(records)} records to sheet "{sheet_name}"')
            
            async with self._write_lock:
                # Ensure sheet exists
                if sheet_name not in self.current_workbook.sheetnames:
                    self.current_workbook.create_sheet(sheet_name)
                    self.sheet_row_counts[sheet_name] = 0
                
                sheet = self.current_workbook[sheet_name]
                
                # Add headers if sheet is empty and headers provided
                start_row = await self._get_next_available_row(sheet_name)
                if start_row == 1 and headers:
                    for col_idx, header in enumerate(headers, 1):
                        sheet.cell(row=1, column=col_idx, value=header)
                    start_row = 2
                
                # Process records in chunks for memory efficiency
                total_written = 0
                current_row = start_row
                
                for chunk_start in range(0, len(records), chunk_size):
                    if client_wants_cancel.is_set():
                        await self._render_media_markdown(
                            '**Append operation cancelled by user**',
                            'append_records',
                            tool_context=tool_context
                        )
                        break
                    
                    chunk_end = min(chunk_start + chunk_size, len(records))
                    chunk_records = records[chunk_start:chunk_end]
                    
                    # Write chunk to sheet
                    for record in chunk_records:
                        for col_idx, value in enumerate(record, 1):
                            sheet.cell(row=current_row, column=col_idx, value=value)
                        current_row += 1
                        total_written += 1
                    
                    # Update progress for large operations
                    if len(records) > chunk_size and chunk_end < len(records):
                        progress = (chunk_end / len(records)) * 100
                        await self._raise_text_delta_event(
                            f'Appending records: {progress:.1f}% complete ({chunk_end}/{len(records)})\n',
                            tool_context=tool_context
                        )
                
                # Update row tracking
                await self._update_sheet_row_count(sheet_name, current_row - 1)
                
                end_row = current_row - 1
                
                self.logger.info(f'Append operation {operation_id} completed: wrote {total_written} records to rows {start_row}-{end_row}')
                
                return json.dumps({
                    'success': True,
                    'message': f'Successfully appended {total_written} records to sheet "{sheet_name}"',
                    'operation_id': operation_id,
                    'agent_id': agent_id,
                    'sheet_name': sheet_name,
                    'records_written': total_written,
                    'start_row': start_row,
                    'end_row': end_row,
                    'headers_added': bool(headers and start_row == 2)
                })
                
        except Exception as e:
            error_msg = f'Error appending records: {str(e)}'
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })

    @json_schema(
        description='Write data to a previously reserved row range - guaranteed conflict-free',
        params={
            'reservation_id': {
                'type': 'string',
                'description': 'ID of the row reservation returned by reserve_rows',
                'required': True
            },
            'records': {
                'type': 'array',
                'description': 'Array of records to write. Each record should be an array of values.',
                'required': True,
                'items': {
                    'type': 'array',
                    'items': {'type': 'string'}
                }
            }
        }
    )
    async def write_to_reservation(self, **kwargs) -> str:
        """
        Write data to a previously reserved row range. This guarantees conflict-free
        writing when multiple agents are working simultaneously.
        
        Args:
            **kwargs: Contains reservation_id and records
            
        Returns:
            str: JSON response with write results
        """
        try:
            if not self.current_workbook:
                return json.dumps({
                    'success': False,
                    'error': 'No workbook is currently loaded. Create or load a workbook first.'
                })
            
            # Validate required fields
            success, message = validate_required_fields(kwargs, ['reservation_id', 'records'])
            if not success:
                return json.dumps({'success': False, 'error': message})
            
            reservation_id = kwargs.get('reservation_id')
            records = kwargs.get('records')
            
            # Check if reservation exists
            if reservation_id not in self.row_reservations:
                return json.dumps({
                    'success': False,
                    'error': f'Reservation {reservation_id} not found or expired'
                })
            
            reservation = self.row_reservations[reservation_id]
            
            if reservation['status'] != 'active':
                return json.dumps({
                    'success': False,
                    'error': f'Reservation {reservation_id} is not active (status: {reservation["status"]})'
                })
            
            # Validate record count fits in reservation
            if len(records) > reservation['row_count']:
                return json.dumps({
                    'success': False,
                    'error': f'Too many records ({len(records)}) for reservation ({reservation["row_count"]} rows)'
                })
            
            sheet_name = reservation['sheet_name']
            start_row = reservation['start_row']
            
            # Get sheet
            sheet = self.current_workbook[sheet_name]
            
            # Write records to reserved rows
            current_row = start_row
            for record in records:
                for col_idx, value in enumerate(record, 1):
                    sheet.cell(row=current_row, column=col_idx, value=value)
                current_row += 1
            
            # Mark reservation as used
            reservation['status'] = 'completed'
            reservation['completed_at'] = datetime.now().isoformat()
            reservation['records_written'] = len(records)
            
            # Update row tracking
            end_row = start_row + len(records) - 1
            await self._update_sheet_row_count(sheet_name, end_row)
            
            self.logger.info(f'Wrote {len(records)} records to reserved rows {start_row}-{end_row} in sheet "{sheet_name}"')
            
            return json.dumps({
                'success': True,
                'message': f'Successfully wrote {len(records)} records to reserved rows',
                'reservation_id': reservation_id,
                'agent_id': reservation['agent_id'],
                'sheet_name': sheet_name,
                'records_written': len(records),
                'start_row': start_row,
                'end_row': end_row
            })
            
        except Exception as e:
            error_msg = f'Error writing to reservation: {str(e)}'
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })

    @json_schema(
        description='Read data from a sheet with optional range specification',
        params={
            'sheet_name': {
                'type': 'string',
                'description': 'Name of the sheet to read from',
                'required': False,
                'default': 'Sheet'
            },
            'start_row': {
                'type': 'integer',
                'description': 'Starting row (1-based). If not provided, reads from row 1.',
                'required': False
            },
            'end_row': {
                'type': 'integer',
                'description': 'Ending row (1-based). If not provided, reads to last row with data.',
                'required': False
            },
            'start_column': {
                'type': 'string',
                'description': 'Starting column (e.g., "A"). If not provided, starts from column A.',
                'required': False
            },
            'end_column': {
                'type': 'string',
                'description': 'Ending column (e.g., "Z"). If not provided, reads to last column with data.',
                'required': False
            },
            'include_headers': {
                'type': 'boolean',
                'description': 'Whether to treat first row as headers',
                'required': False,
                'default': False
            },
            'max_rows': {
                'type': 'integer',
                'description': 'Maximum number of rows to read (for large sheets)',
                'required': False,
                'default': 10000
            }
        }
    )
    async def read_sheet_data(self, **kwargs) -> str:
        """
        Read data from a sheet with optional range specification.
        For large sheets, data may be cached if it exceeds token limits.
        
        Args:
            **kwargs: Contains sheet_name and range parameters
            
        Returns:
            str: JSON response with sheet data or cache key
        """
        try:
            if not self.current_workbook:
                return json.dumps({
                    'success': False,
                    'error': 'No workbook is currently loaded. Create or load a workbook first.'
                })
            
            sheet_name = kwargs.get('sheet_name', 'Sheet')
            start_row = kwargs.get('start_row', 1)
            end_row = kwargs.get('end_row')
            start_column = kwargs.get('start_column', 'A')
            end_column = kwargs.get('end_column')
            include_headers = kwargs.get('include_headers', False)
            max_rows = kwargs.get('max_rows', 10000)
            
            # Check if sheet exists
            if sheet_name not in self.current_workbook.sheetnames:
                return json.dumps({
                    'success': False,
                    'error': f'Sheet "{sheet_name}" does not exist'
                })
            
            sheet = self.current_workbook[sheet_name]
            
            # Determine actual range
            if not end_row:
                end_row = min(sheet.max_row, start_row + max_rows - 1)
            
            if not end_column:
                end_column = get_column_letter(sheet.max_column)
            
            # Convert column letters to indices
            start_col_idx = column_index_from_string(start_column)
            end_col_idx = column_index_from_string(end_column)
            
            # Read data
            data = []
            headers = []
            
            for row_idx, row in enumerate(sheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_col_idx,
                max_col=end_col_idx,
                values_only=True
            )):
                if row_idx == 0 and include_headers:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                else:
                    data.append([str(cell) if cell is not None else '' for cell in row])
            
            # Prepare response
            result = {
                'success': True,
                'sheet_name': sheet_name,
                'range': f'{start_column}{start_row}:{end_column}{end_row}',
                'rows_read': len(data),
                'columns_read': end_col_idx - start_col_idx + 1,
            }
            
            if include_headers:
                result['headers'] = headers
            
            # Check if response is too big for direct return
            response_data = {
                'data': data,
                'headers': headers if include_headers else None
            }
            response_json = json.dumps(response_data)
            
            if self._is_response_too_big(response_json):
                # Cache the data and return cache key
                cache_key = f'excel_data_{int(time.time() * 1000)}_{uuid.uuid4().hex[:8]}'
                self.tool_cache.set(cache_key, response_data, expire=3600)  # 1 hour
                
                result['cached'] = True
                result['cache_key'] = cache_key
                result['message'] = 'Data cached due to size. Use cache key to access full dataset.'
                
                self.logger.info(f'Cached sheet data with key {cache_key}: {len(data)} rows')
            else:
                result['data'] = data
                if include_headers:
                    result['headers'] = headers
            
            return json.dumps(result)
            
        except Exception as e:
            error_msg = f'Error reading sheet data: {str(e)}'
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })

    @json_schema(
        description='Load cached sheet data using a cache key',
        params={
            'cache_key': {
                'type': 'string',
                'description': 'Cache key returned by read_sheet_data when data was too large',
                'required': True
            }
        }
    )
    async def load_cached_data(self, **kwargs) -> str:
        """
        Load cached sheet data using a cache key.
        
        Args:
            **kwargs: Contains cache_key
            
        Returns:
            str: JSON response with cached data
        """
        try:
            # Validate required fields
            success, message = validate_required_fields(kwargs, ['cache_key'])
            if not success:
                return json.dumps({'success': False, 'error': message})
            
            cache_key = kwargs.get('cache_key')
            
            # Retrieve cached data
            cached_data = self.tool_cache.get(cache_key)
            
            if cached_data is None:
                return json.dumps({
                    'success': False,
                    'error': f'No cached data found for key {cache_key} (may have expired)'
                })
            
            return json.dumps({
                'success': True,
                'message': 'Cached data retrieved successfully',
                'cache_key': cache_key,
                **cached_data
            })
            
        except Exception as e:
            error_msg = f'Error loading cached data: {str(e)}'
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })



    @json_schema(
        description='Get the next available row number in a sheet for coordination between agents',
        params={
            'sheet_name': {
                'type': 'string',
                'description': 'Name of the sheet to check',
                'required': False,
                'default': 'Sheet'
            }
        }
    )
    async def get_next_available_row(self, **kwargs) -> str:
        """
        Get the next available row number in a sheet. Useful for agents
        to coordinate their writing operations.
        
        Args:
            **kwargs: Contains optional sheet_name
            
        Returns:
            str: JSON response with row information
        """
        try:
            if not self.current_workbook:
                return json.dumps({
                    'success': False,
                    'error': 'No workbook is currently loaded. Create or load a workbook first.'
                })
            
            sheet_name = kwargs.get('sheet_name', 'Sheet')
            
            # Ensure sheet exists
            if sheet_name not in self.current_workbook.sheetnames:
                return json.dumps({
                    'success': True,
                    'sheet_name': sheet_name,
                    'next_available_row': 1,
                    'sheet_exists': False
                })
            
            next_row = await self._get_next_available_row(sheet_name)
            
            # Account for active reservations
            reserved_rows = sum(
                res['row_count'] for res in self.row_reservations.values()
                if res['sheet_name'] == sheet_name and res['status'] == 'active'
            )
            
            return json.dumps({
                'success': True,
                'sheet_name': sheet_name,
                'next_available_row': next_row,
                'reserved_rows': reserved_rows,
                'next_safe_row': next_row + reserved_rows,
                'sheet_exists': True
            })
            
        except Exception as e:
            error_msg = f'Error getting next available row: {str(e)}'
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })


    @json_schema(
        description='Get operation status and statistics for monitoring large operations',
        params={
            'operation_id': {
                'type': 'string',
                'description': 'Operation ID returned by write operations',
                'required': False
            }
        }
    )
    async def get_operation_status(self, **kwargs) -> str:
        """
        Get status and statistics for operations. Useful for monitoring
        large write operations and debugging.
        
        Args:
            **kwargs: Contains optional operation_id
            
        Returns:
            str: JSON response with operation status
        """
        try:
            operation_id = kwargs.get('operation_id')
            
            result = {
                'success': True,
                'current_workbook': {
                    'loaded': self.current_workbook is not None,
                    'path': self.current_workbook_path,
                    'sheets': len(self.current_workbook.sheetnames) if self.current_workbook else 0,
                    'metadata': self.current_workbook_metadata
                },
                'row_tracking': dict(self.sheet_row_counts),
                'active_reservations': {
                    res_id: {
                        'agent_id': res['agent_id'],
                        'sheet_name': res['sheet_name'],
                        'row_range': f"{res['start_row']}-{res['end_row']}",
                        'status': res['status']
                    }
                    for res_id, res in self.row_reservations.items()
                    if res['status'] == 'active'
                },
                'total_reservations': len(self.row_reservations)
            }
            
            if operation_id and operation_id in self.active_operations:
                result['operation_details'] = self.active_operations[operation_id]
            
            return json.dumps(result)
            
        except Exception as e:
            error_msg = f'Error getting operation status: {str(e)}'
            self.logger.error(error_msg)
            return json.dumps({
                'success': False,
                'error': error_msg
            })


# Register the toolset
Toolset.register(ExcelTools, required_tools=['WorkspaceTools'])